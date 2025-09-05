from django.contrib import messages
from django.shortcuts import redirect, render
from django.views.decorators.csrf import csrf_exempt
from django_mailbox.form import EmailConfigForm
from rest_framework import status
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.views import APIView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.core.cache import cache
from .tasks import get_sparkai_response_task
import json
import time
import os
import uuid
from .models import EmailConfig
import datetime
from django.db import transaction
from .models import BatchTask, TaskItem
from .serializers import EmailConfigSerializer,SendMailSerializer

from .utils import send_email_with_attachment
import logging

logger = logging.getLogger(__name__)

class BaseModelViewSet(viewsets.ModelViewSet):
    '''
    基础视图集，可在此添加通用逻辑
    '''
    pass


# 自动生成的模型视图集

class EmailConfigViewSet(BaseModelViewSet):
    queryset = EmailConfig.objects.all()
    serializer_class = EmailConfigSerializer


@csrf_exempt  # 临时保留以便调试
def index(request):
    try:
        email_config = EmailConfig.objects.latest('updated_at')
        logger.info(f"现有配置: {email_config.__dict__}")
    except EmailConfig.DoesNotExist:
        email_config = None
        logger.info("无现有配置")

    if request.method == 'POST':
        logger.info(f"收到POST数据: {request.POST}")
        form = EmailConfigForm(request.POST, instance=email_config)
        
        if form.is_valid():
            instance = form.save()
            logger.info(f"保存成功! 新实例ID: {instance.id}")
            messages.success(request, "配置已保存")
            return redirect('django_mailbox:email')
        else:
            logger.error(f"表单无效: {form.errors}")
    else:
        form = EmailConfigForm(instance=email_config)

    return render(request, "django_mailbox/email.html", {'form': form})


class SendMailAPIView(APIView):
    @csrf_exempt
    def post(self, request):
        # 打印当前路径
        serializer = SendMailSerializer(data=request.data)
        if serializer.is_valid():
            subject = serializer.validated_data['subject']
            content = serializer.validated_data['content']
            filepath = serializer.validated_data.get('filepath')
            # 打印当前路径
            if filepath:
                import os
                print('邮件附件绝对路径:', os.path.abspath(filepath))
            try:
                send_email_with_attachment(subject, content, filepath)
                return Response({"detail": "邮件发送成功"})
            except Exception as e:
                return Response({"detail": f"邮件发送失败: {e}"}, status=status.HTTP_400_BAD_REQUEST)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




def get_short_uuid(uuid_obj):
    """获取UUID的短格式（前8位）"""
    return str(uuid_obj)[:8]


def timestamp_to_human(timestamp):
    # 转换为UTC时间
    utc_time = datetime.datetime.fromtimestamp(timestamp, datetime.timezone.utc)
    # 转换为北京时间 (UTC+8)
    beijing_time = utc_time.astimezone(datetime.timezone(datetime.timedelta(hours=8)))
    
    # 格式化输出
    formatted_time = beijing_time.strftime("%Y年%m月%d日 %H:%M:%S")
    
    return formatted_time

def batch_tasks_view(request):
    """
    批量任务提交界面 - 支持文件上传
    """
    if request.method == 'POST':
        questions = []
        
        # 处理文件上传
        uploaded_file = request.FILES.get('questions_file')
        if uploaded_file:
            try:
                # 读取上传的文件内容
                content = uploaded_file.read().decode('utf-8')
                # 按行分割，过滤空行
                file_questions = [line.strip() for line in content.split('\n') if line.strip()]
                questions.extend(file_questions)
            except Exception as e:
                messages.error(request, f'文件读取失败: {str(e)}')
                return redirect('django_mailbox:batch_tasks')
        
        # 处理手动输入的问题
        for key in request.POST:
            if key.startswith('question_') and request.POST[key].strip():
                questions.append(request.POST[key].strip())
        
        if not questions:
            messages.error(request, '请至少输入一个有效的问题或上传文件')
            return redirect('django_mailbox:batch_tasks')
        
        # 限制最大问题数量，避免性能问题
        if len(questions) > 100:
            messages.warning(request, f'问题数量超过100个，将只处理前100个问题')
            questions = questions[:100]
        
        # 使用事务创建批量任务记录
        with transaction.atomic():
            batch_task = BatchTask.objects.create(
                user=request.user if request.user.is_authenticated else None,
                total_questions=len(questions)
            )
            
            task_items = []
            for question in questions:
                task_item = TaskItem(
                    batch_task=batch_task,
                    question=question
                )
                task_items.append(task_item)
            
            # 批量创建任务项
            TaskItem.objects.bulk_create(task_items)
            
            # 启动异步任务
            for task_item in task_items:
                get_sparkai_response_task.delay(task_item.question, str(task_item.id))
        
        messages.success(request, f'已提交 {len(questions)} 个任务，请查看任务状态')
        return redirect('django_mailbox:task_status', batch_id=batch_task.id)
    
    # GET请求显示空表单
    return render(request, "django_mailbox/batch_tasks.html")

def task_status(request, batch_id=None):
    """
    任务状态查看页面 - 添加分页功能
    """
    # 如果没有指定batch_id，尝试获取最新的任务
    if not batch_id:
        latest_task = BatchTask.objects.order_by('-submitted_at').first()
        if latest_task:
            return redirect('django_mailbox:task_status', batch_id=latest_task.id)
        else:
            messages.info(request, '暂无任务记录')
            return redirect('django_mailbox:batch_tasks')
    
    try:
        batch_task = BatchTask.objects.get(id=batch_id)
    except BatchTask.DoesNotExist:
        messages.error(request, '任务不存在')
        return redirect('django_mailbox:batch_tasks')
    
    # 获取任务项并更新统计信息
    task_items = batch_task.items.all()
    
    # 实时更新统计信息
    completed_count = task_items.filter(status='success').count()
    error_count = task_items.filter(status='error').count()
    processing_count = task_items.filter(status='processing').count()
    
    # 更新批量任务的统计信息
    if (batch_task.completed_count != completed_count or 
        batch_task.error_count != error_count):
        batch_task.completed_count = completed_count
        batch_task.error_count = error_count
        batch_task.save(update_fields=['completed_count', 'error_count'])
    
    # 准备结果数据
    results = []
    for item in task_items:
        result_data = cache.get(f'sparkai_result_{item.id}')
        
        # 如果缓存中有更新的结果，更新数据库
        if result_data and result_data.get('status') != item.status:
            item.status = result_data['status']
            item.result = result_data.get('result')
            item.error_message = result_data.get('error')
            if result_data.get('completed_at'):
                item.completed_at = datetime.datetime.fromtimestamp(
                    result_data['completed_at'], datetime.timezone.utc
                )
            item.save()
        
        results.append({
            'question': item.question,
            'task_id': item.id,
            'status': item.status,
            'result': item.result,
            'error': item.error_message,
            'completed_at': item.completed_at.strftime("%Y年%m月%d日 %H:%M:%S") if item.completed_at else None
        })
    
    # 添加分页功能
    per_page = int(request.GET.get('per_page', 10))
    per_page = min(per_page, 20)  # 限制最大每页20条

    paginator = Paginator(results, per_page)
    page_number = request.GET.get('page', 1)
    
    try:
        page_results = paginator.page(page_number)
    except PageNotAnInteger:
        page_results = paginator.page(1)
    except EmptyPage:
        page_results = paginator.page(paginator.num_pages)
    
    context = {
        'batch_task': batch_task,
        'results': page_results,
        'total_count': batch_task.total_questions,
        'completed_count': completed_count,
        'processing_count': processing_count,
        'error_count': error_count,
        'progress_percent': batch_task.progress_percent(),
        'submitted_at': batch_task.submitted_at,
        'all_batch_tasks': BatchTask.objects.order_by('-submitted_at')[:10]  # 显示最近10个任务
    }
    
    return render(request, "django_mailbox/task_status.html", context)

def task_history(request):
    """
    任务历史记录页面
    """
    batch_tasks = BatchTask.objects.order_by('-submitted_at')
    
    # 添加分页
    paginator = Paginator(batch_tasks, 10)  # 每页10个任务
    page_number = request.GET.get('page', 1)
    
    try:
        page_tasks = paginator.page(page_number)
    except PageNotAnInteger:
        page_tasks = paginator.page(1)
    except EmptyPage:
        page_tasks = paginator.page(paginator.num_pages)
    
    return render(request, "django_mailbox/task_history.html", {'batch_tasks': page_tasks})






# views.py
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse
from django.utils import timezone
from django.shortcuts import get_object_or_404

def export_tasks_excel(request, batch_id):
    """
    导出任务结果到Excel
    """
    batch_task = get_object_or_404(BatchTask, id=batch_id)
    task_items = batch_task.items.all().order_by('created_at')
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"批量任务结果_{get_short_uuid(batch_id)}"
    
    # 设置标题行
    headers = ['序号', '问题内容', '状态', 'AI回复', '错误信息', '完成时间', '创建时间']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    
    # 填充数据
    for row_num, item in enumerate(task_items, 2):
        ws.cell(row=row_num, column=1).value = row_num - 1  # 序号
        ws.cell(row=row_num, column=2).value = item.question
        ws.cell(row=row_num, column=3).value = item.get_status_display()
        
        # 状态颜色
        status_cell = ws.cell(row=row_num, column=3)
        if item.status == 'success':
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif item.status == 'error':
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        ws.cell(row=row_num, column=4).value = item.result
        ws.cell(row=row_num, column=5).value = item.error_message
        
        # 时间格式化
        if item.completed_at:
            ws.cell(row=row_num, column=6).value = item.completed_at.strftime("%Y-%m-%d %H:%M:%S")
        if item.created_at:
            ws.cell(row=row_num, column=7).value = item.created_at.strftime("%Y-%m-%d %H:%M:%S")
    
    # 调整列宽
    column_widths = [8, 50, 12, 60, 40, 20, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 设置自动换行
    for row in ws.iter_rows(min_row=2, max_row=len(task_items)+1, min_col=2, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 创建响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"批量任务导出_{get_short_uuid(batch_id)}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

def export_tasks_summary_excel(request, batch_id):
    """
    导出任务统计摘要
    """
    batch_task = get_object_or_404(BatchTask, id=batch_id)
    task_items = batch_task.items.all()
    
    # 统计信息
    total_count = task_items.count()
    success_count = task_items.filter(status='success').count()
    error_count = task_items.filter(status='error').count()
    processing_count = task_items.filter(status='processing').count()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "任务统计摘要"
    
    # 摘要信息
    summary_data = [
        ['批量任务ID', str(batch_task.id)],
        ['提交时间', batch_task.submitted_at.strftime("%Y-%m-%d %H:%M:%S")],
        ['总问题数', total_count],
        ['成功数量', success_count],
        ['失败数量', error_count],
        ['处理中数量', processing_count],
        ['成功率', f"{success_count/total_count*100:.2f}%" if total_count > 0 else "0%"],
        ['完成进度', f"{batch_task.progress_percent()}%"]
    ]
    
    for row_num, (label, value) in enumerate(summary_data, 1):
        ws.cell(row=row_num, column=1).value = label
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        ws.cell(row=row_num, column=2).value = value
    
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"任务统计摘要_{get_short_uuid(batch_id)}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response